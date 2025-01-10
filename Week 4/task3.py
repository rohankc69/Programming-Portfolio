import tkinter as tk
from tkinter import ttk, messagebox
from openpyxl import Workbook, load_workbook
import os

root = tk.Tk()
root.title("Data Store System")
root.geometry("700x600")
FILE_NAME = "Task3.xlsx"

def initialize_excel():
    if not os.path.exists(FILE_NAME):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["Name", "Age"])
        workbook.save(FILE_NAME)
        print("No file found, created a new Excel file.")
    else:
        print("File found.")

initialize_excel()

ui_frame = tk.Frame(root)
ui_frame.grid(row=1, column=0, sticky='nsew')

root.grid_rowconfigure(1, weight=1)  
root.grid_columnconfigure(0, weight=1)  


tk.Label(ui_frame, text="Name").grid(row=0, column=0, sticky='ew')
entry_name = tk.Entry(ui_frame, width=40)  
entry_name.grid(row=0, column=1, sticky='ew')

tk.Label(ui_frame, text="Age").grid(row=1, column=0, sticky='ew')
enter_age = tk.Entry(ui_frame, width=40)  
enter_age.grid(row=1, column=1, sticky='ew')


table = ttk.Treeview(ui_frame, columns=("Name", "Age"), show="headings")
table.heading("Name", text="Name ")
table.heading("Age", text="Age")
table.grid(row=3, column=0, columnspan=3, sticky='nsew')


ui_frame.grid_rowconfigure(3, weight=1)  
ui_frame.grid_columnconfigure(0, weight=1)  


def load_data():
    workbook = load_workbook(FILE_NAME)
    worksheet = workbook.active
    for row in worksheet.iter_rows(min_row=2, values_only=True):  
        table.insert("", "end", values=row)

load_data()  


def save_data():
    name = entry_name.get()
    age = enter_age.get()

   
    table.insert("", "end", values=(name, age))
    
  
    workbook = load_workbook(FILE_NAME)
    worksheet = workbook.active
    worksheet.append([name, age])  
    workbook.save(FILE_NAME)  
    
    
    entry_name.delete(0, tk.END)
    enter_age.delete(0, tk.END)
    
    print("Your data has been saved.")


def update_data():
    selected_item = table.selection()
    if not selected_item:
        messagebox.showwarning("Update Error", "Please select an entry to update.")
        return

    name = entry_name.get()
    age = enter_age.get()

    if not name or not age:
        messagebox.showwarning("Input Error", "Please provide both Name and Age.")
        return


    table.item(selected_item, values=(name, age))

  
    workbook = load_workbook(FILE_NAME)
    worksheet = workbook.active
    index = table.index(selected_item) + 2  
    worksheet.cell(row=index, column=1, value=name)
    worksheet.cell(row=index, column=2, value=age)
    workbook.save(FILE_NAME)


    entry_name.delete(0, tk.END)
    enter_age.delete(0, tk.END)
    
    print("Your data has been updated.")


def delete_data():
    selected_item = table.selection()
    if not selected_item:
        messagebox.showwarning("Delete Error", "Please select an entry to delete.")
        return

    item_values = table.item(selected_item, 'values')  
    name, age = item_values  

    table.delete(selected_item)


    workbook = load_workbook(FILE_NAME)
    worksheet = workbook.active


    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if row[0] == name and row[1] == age:
            worksheet.delete_rows(row[0].row)  
            break

    workbook.save(FILE_NAME)

    print("Your data has been deleted.")


tk.Button(ui_frame, text="Save", command=save_data, width=10).grid(row=2, column=0, sticky='ew') 


tk.Button(ui_frame, text="Update", command=update_data, width=10).grid(row=2, column=1, sticky='ew')


tk.Button(ui_frame, text="Delete", command=delete_data, width=10).grid(row=2, column=2, sticky='ew')


def on_item_double_click(event):
    selected_item = table.selection()
    if selected_item:
        item_values = table.item(selected_item, 'values')
        entry_name.delete(0, tk.END)
        entry_name.insert(0, item_values[0])
        enter_age.delete(0, tk.END)
        enter_age.insert(0, item_values[1])

table.bind("<Double-1>", on_item_double_click)

root.mainloop()
